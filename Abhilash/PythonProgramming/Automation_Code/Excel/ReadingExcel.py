import openpyxl

#File -- workbook -- sheet--- row---cell

# file_path = "E:\Practice files\sales1.xlsx"
# workbook= openpyxl.load_workbook(file_path)
#
# # get active sheet from excel
# sheet = workbook.active
#
# # if we have multiple sheets we can mention sheet name
# # sheet= workbook["sales"]

# rows = sheet.max_row
# cols = sheet.max_column

#read all the rows and columns from excel sheet
# for r in range(1,rows+1):
#     for c in range(1,cols+1):
#         print(sheet.cell(r,c).value,end='        ')
#     print()


# Write the data in Excel

#File -- workbook -- sheet--- write the date in row and coulmns --- save

file_path = r"E:\Practice files\Writemultipledata.xlsx"
workbook= openpyxl.load_workbook(file_path)

# get active sheet from excel
sheet = workbook.active
# if we have multiple sheets we can mention sheet name
# sheet= workbook["sales"]

#write only single date in the sheet
# for r in range(1,6):
#     for c in range(1,4):
#         sheet.cell(r,c).value="Sales"
#
# workbook.save(file_path)

#Write multiple date in the sheet

sheet.cell(1,1).value='Region'
sheet.cell(1,2).value='SalesPerson'

sheet.cell(2,1).value='South'
sheet.cell(3,1).value='East'
sheet.cell(4,1).value='North'
sheet.cell(5,1).value='West'

sheet.cell(2,2).value='A'
sheet.cell(3,2).value='B'
sheet.cell(4,2).value='C'
sheet.cell(5,2).value='D'

workbook.save(file_path)
