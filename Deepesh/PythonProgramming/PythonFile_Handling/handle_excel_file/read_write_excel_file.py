# install package with command
# -> pip install openpyxl
import openpyxl

def read_excel_file(file_path, sheet_name, cell_name):
    wb = openpyxl.load_workbook(file_path)
    sheet = wb[sheet_name]
    cell = sheet[cell_name]
    print(cell.value)


# read_excel_file("users_data.xlsx", "Sheet1", "A1") # Pune
# read_excel_file("users_data.xlsx", "Sheet2", "A1") # India
# read_excel_file("users_data.xlsx", "Sheet3", "A1") # Red

# for i in range(1, 5):
#     read_excel_file("users_data.xlsx", "Sheet1", f"A{i}")
"""
Pune
Mumbai
Bangalore
Kolkata

"""
#
# for i in range(1, 4):
#     for j in range(1, 5):
#         read_excel_file("users_data.xlsx", f"Sheet{i}", f"A{j}")
#
#     print("_"*50)
"""
Pune
Mumbai
Bangalore
Kolkata
__________________________________________________
India
China
USA
Russia
__________________________________________________
Red
Green
Black
Blue
__________________________________________________
"""


def read_all_row_colums_values(file_path, sheet_name):
    wb = openpyxl.load_workbook(file_path)
    sheet = wb[sheet_name]
    max_row = sheet.max_row
    max_coln = sheet.max_column
    print(max_row, max_coln)
    for i in range(1, max_row+1):
        for j in range(1, max_coln+1):
            cell = sheet.cell(i, j)
            print(cell.value, end='  |  ')
        print()


#read_all_row_colums_values("users_data.xlsx", sheet_name='Sheet3')

