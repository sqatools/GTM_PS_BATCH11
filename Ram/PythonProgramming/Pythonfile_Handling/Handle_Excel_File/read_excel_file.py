# install package with command
# -> pip install openpyxl
import openpyxl
def read_excel_data(filepath, sheet_name, cell_name):
    wb = openpyxl.load_workbook(filepath)
    sheet = wb[sheet_name]
    cell = sheet[cell_name]
    print(cell.value)


read_excel_data(r"D:\Python Practice\read_data.xlsx", "Sheet1", "A1")  # Red

for i in range(1, 6):
    read_excel_data(r"D:\Python Practice\read_data.xlsx", "Sheet1", f"A{i}")
print("_"*50)
"""
Red
Blue    
Green
Yellow
White
"""

# Program to data from two Sheets
for i in range(1,3):
    for j in range(1, 5):
        read_excel_data(r"D:\Python Practice\read_data.xlsx", f"Sheet{i}", f"A{j}")
    print("_"*50)
"""
Red
Blue
Green
Yellow
__________________________________________________
Sachin
Kohli
Sehwag
Dravid
__________________________________________________
"""
print("*"*50)
def read_two_columns_from_sheet(file_path, sheet_name, cell1_name, cell2_name):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]
    for i in range(1, 7):
        val1 = sheet[f'{cell1_name}{i}'].value
        val2 = sheet[f'{cell2_name}{i}'].value
        print(f"{val1}, {val2}")

"""
Sachin, 20
Kohli, 18
Sehwag, 44
Dravid, 24
Yuvaraj, 12
Rohith, 46
"""
#read_two_columns_from_sheet(r"D:\Python Practice\read_data.xlsx", "Sheet2", "A", "B")

print("_"*50)
# Program to print Total No of rows and columns from one sheet
print("_"*50)
def read_all_row_columns_values(file_path, sheet_name):
    wb = openpyxl.load_workbook(file_path)
    sheet = wb[sheet_name]
    max_row = sheet.max_row
    max_coln = sheet.max_column
    print(f"Row: {max_row}", f" Col: {max_coln}")
    for i in range(1, max_row+1):
        for j in range(1, max_coln+1):
            cell = sheet.cell(i, j)
            print(cell.value, end='  |  ')
        print()


#read_all_row_columns_values(r"D:\Python Practice\read_data.xlsx", sheet_name='Sheet1')
"""
Row: 8  Col: 3
Red  |  10  |  1  |  
Blue  |  20  |  2  |  
Green  |  30  |  3  |  
Yellow  |  40  |  4  |  
White  |  50  |  5  |  
Pink  |  60  |  6  |  
Black  |  70  |  7  |  
Orange  |  80  |  8  |
"""
# # Program to print Total No of rows and columns from multiple sheets
def read_rows_columns_from_multiple_sheets(file_path, sheet_names):
    workbook = openpyxl.load_workbook(file_path)
    for sheet_name in sheet_names:
        sheet = workbook[sheet_name]
        max_row = sheet.max_row
        max_column = sheet.max_column
        print(f"{sheet_name}")
        print("_" * 50)
        for i in range(1, max_row + 1):
            for j in range(1, max_column + 1):
                cell = sheet.cell(i, j)
                print(cell.value, end='  |  ')
            print()

sheet_names = ['Sheet1', 'Sheet2']
#read_rows_columns_from_multiple_sheets(r"D:\Python Practice\read_data.xlsx", sheet_names)
"""
Sheet1
__________________________________________________
Red  |  10  |  1  |  
Blue  |  20  |  2  |  
Green  |  30  |  3  |  
Yellow  |  40  |  4  |  
White  |  50  |  5  |  
Pink  |  60  |  6  |  
Black  |  70  |  7  |  
Orange  |  80  |  8  |  
Sheet2
__________________________________________________
Sachin  |  20  |  
Kohli  |  18  |  
Sehwag  |  44  |  
Dravid  |  24  |  
Yuvaraj  |  12  |  
Rohith  |  46  |  
"""
################## OR ######################

# Program to print Total No of rows and columns from multiple sheets
def read_all_row_columns_values1(file_path, sheet_name1, sheet_name2):
    wb = openpyxl.load_workbook(file_path)
    sheet1 = wb[sheet_name1]
    sheet2 = wb[sheet_name2]
    max_row = sheet1.max_row
    max_column = sheet1.max_column
    print("Sheet1")
    print(f"Row: {max_row}", f" Col: {max_column}")
    for i in range(1, max_row+1):
        for j in range(1, max_column+1):
            cell1 = sheet1.cell(i, j)
            print(cell1.value, end='  |  ')
        print()
    print("_"*50)
    print("Sheet2")
    for k in range(1, 7):
        for l in range(1, 3):
            cell2 = sheet2.cell(k, l)
            print(cell2.value, end='  |  ')
        print()


#read_all_row_columns_values1(r"D:\Python Practice\read_data.xlsx", sheet_name1='Sheet1', sheet_name2='Sheet2')
"""
Sheet1
Row: 8  Col: 3
Red  |  10  |  1  |  
Blue  |  20  |  2  |  
Green  |  30  |  3  |  
Yellow  |  40  |  4  |  
White  |  50  |  5  |  
Pink  |  60  |  6  |  
Black  |  70  |  7  |  
Orange  |  80  |  8  |  
__________________________________________________
Sheet2
Sachin  |  20  |  
Kohli  |  18  |  
Sehwag  |  44  |  
Dravid  |  24  |  
Yuvaraj  |  12  |  
Rohith  |  46  |
"""

# Program to print alternate rows and columns from multiple sheets
def read_alternate_rows_columns_from_multiple_sheets(file_path, sheet_names):
    workbook = openpyxl.load_workbook(file_path)
    for sheet_name in sheet_names:
        sheet = workbook[sheet_name]
        max_row = sheet.max_row
        max_column = sheet.max_column
        print(f"{sheet_name}")
        print("_" * 50)
        for i in range(1, max_row + 1, 2):
            for j in range(1, max_column + 1):
                cell = sheet.cell(i, j)
                print(cell.value, end='  |  ')
            print()

sheet_names = ['Sheet1', 'Sheet2']
read_alternate_rows_columns_from_multiple_sheets(r"D:\Python Practice\read_data.xlsx", sheet_names)